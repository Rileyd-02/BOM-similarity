import streamlit as st
import pandas as pd
import difflib
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# --- Streamlit Config ---
st.set_page_config(layout="wide", page_title="SAP & PLM BOM Validation")

st.title("📊 SAP vs PLM BOM Validation Tool")
st.write("Upload SAP and PLM Excel files to compare material numbers and consumption quantities.")

# --- File Uploads ---
sap_file = st.file_uploader("Upload SAP File", type=["xlsx", "xls"])
plm_file = st.file_uploader("Upload PLM File", type=["xlsx", "xls"])

if sap_file and plm_file:
    try:
        sap = pd.read_excel(sap_file, sheet_name="SAP")
        plm = pd.read_excel(plm_file, sheet_name="PLM")

        # --- Step 1: Direct Material Match ---
        direct_matches = pd.merge(
            plm, sap, on="Material", how="inner", suffixes=("_PLM", "_SAP")
        )
        missing_in_sap = plm[~plm["Material"].isin(sap["Material"])]
        missing_in_plm = sap[~sap["Material"].isin(plm["Material"])]

        # --- Add consumption comparison for direct matches ---
        if not direct_matches.empty:
            direct_matches["ConsumptionDiff"] = (
                direct_matches["Qty(Cons.)"] - direct_matches["Comp.Qty."]
            )
            # Add DifferenceFlag
            direct_matches["DifferenceFlag"] = direct_matches.apply(
                lambda r: "SAP Higher" if r["Comp.Qty."] > r["Qty(Cons.)"] else "OK",
                axis=1
            )
            # Sort by largest absolute difference
            direct_matches = direct_matches.reindex(
                direct_matches["ConsumptionDiff"].abs().sort_values(ascending=False).index
            )

        # --- Step 2: Build Combined Column for PLM ---
        plm["Combined"] = (
            plm["Material"].astype(str).str.strip() + " " +
            plm["Vendor Reference"].astype(str).str.strip() + " " +
            plm["Color Reference"].astype(str).str.strip() + " " +
            plm["Color Name"].astype(str).str.strip()
        )

        # --- Step 3: Fuzzy Matching (≥70%) ---
        fuzzy_matches = []
        for _, row in direct_matches.iterrows():
            combined_val = (
                str(row["Material"]).strip() + " " +
                str(row["Vendor Reference_PLM"]).strip() + " " +
                str(row["Color Reference"]).strip() + " " +
                str(row["Color Name"]).strip()
            )
            best_match = difflib.get_close_matches(
                combined_val, sap["Material Description"], n=1, cutoff=0.7
            )
            if best_match:
                sap_row = sap[sap["Material Description"] == best_match[0]].iloc[0]
                plm_val = row.get("Qty(Cons.)", 0)
                sap_val = sap_row.get("Comp.Qty.", 0)
                fuzzy_matches.append({
                    "Material": row["Material"],
                    "Combined_PLMMeta": combined_val,
                    "MaterialDescription_SAP": best_match[0],
                    "Qty(Cons.)_PLM": plm_val,
                    "Comp.Qty._SAP": sap_val,
                    "ConsumptionDiff": plm_val - sap_val,
                    "DifferenceFlag": "SAP Higher" if sap_val > plm_val else "OK",
                    "Vendor Reference_PLM": row.get("Vendor Reference_PLM", ""),
                    "Vendor Reference_SAP": sap_row.get("Vendor Reference", ""),
                    "Color Reference_PLM": row.get("Color Reference", ""),
                    "Comp. Colour_SAP": sap_row.get("Comp. Colour", "")
                })

        fuzzy_df = pd.DataFrame(fuzzy_matches)
        if not fuzzy_df.empty:
            fuzzy_df = fuzzy_df.reindex(
                fuzzy_df["ConsumptionDiff"].abs().sort_values(ascending=False).index
            )

        # --- Save Results to Excel ---
        output = BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            direct_matches.to_excel(writer, sheet_name="Direct_Matches", index=False)
            missing_in_sap.to_excel(writer, sheet_name="PLM_Not_in_SAP", index=False)
            missing_in_plm.to_excel(writer, sheet_name="SAP_Not_in_PLM", index=False)
            fuzzy_df.to_excel(writer, sheet_name="70%_or_more_Matches", index=False)

        output.seek(0)

        # --- Apply Conditional Formatting ---
        wb = load_workbook(output)

        def apply_coloring(ws, headers, plm_col_name, sap_col_name, diff_col_name, flag_col_name):
            plm_col = headers.index(plm_col_name) + 1
            sap_col = headers.index(sap_col_name) + 1
            diff_col = headers.index(diff_col_name) + 1
            flag_col = headers.index(flag_col_name) + 1

            green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
            red_fill = PatternFill(start_color="FF6666", end_color="FF6666", fill_type="solid")

            for row in range(2, ws.max_row + 1):
                plm_val = ws.cell(row=row, column=plm_col).value
                sap_val = ws.cell(row=row, column=sap_col).value

                try:
                    plm_val = float(plm_val or 0)
                    sap_val = float(sap_val or 0)
                except ValueError:
                    continue  # skip rows with invalid data

                if sap_val > plm_val:  # ❌ SAP higher → Red
                    fill = red_fill
                else:  # ✅ PLM ≥ SAP → Green
                    fill = green_fill

                ws.cell(row=row, column=plm_col).fill = fill
                ws.cell(row=row, column=sap_col).fill = fill
                ws.cell(row=row, column=diff_col).fill = fill
                ws.cell(row=row, column=flag_col).fill = fill

        if "Direct_Matches" in wb.sheetnames:
            ws = wb["Direct_Matches"]
            headers = [cell.value for cell in ws[1]]
            apply_coloring(ws, headers, "Qty(Cons.)", "Comp.Qty.", "ConsumptionDiff", "DifferenceFlag")

        if "70%_or_more_Matches" in wb.sheetnames:
            ws = wb["70%_or_more_Matches"]
            headers = [cell.value for cell in ws[1]]
            apply_coloring(ws, headers, "Qty(Cons.)_PLM", "Comp.Qty._SAP", "ConsumptionDiff", "DifferenceFlag")

        # Save updated workbook
        final_output = BytesIO()
        wb.save(final_output)
        final_output.seek(0)

        # --- Summary Section ---
        st.subheader("📊 Summary")
        col1, col2, col3 = st.columns(3)

        total_matches = len(direct_matches)
        sap_higher_count = (direct_matches["DifferenceFlag"] == "SAP Higher").sum()
        ok_count = (direct_matches["DifferenceFlag"] == "OK").sum()

        col1.metric("Total Direct Matches", total_matches)
        col2.metric("❌ SAP Higher", sap_higher_count)
        col3.metric("✅ OK", ok_count)

        # --- Download Button ---
        st.success("✅ Comparison complete! Download the results below.")
        st.download_button(
            label="📥 Download Comparison Report",
            data=final_output,
            file_name="comparison_results.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

        # --- Preview ---
        st.subheader("🔍 Preview of Results")
        tab1, tab2, tab3, tab4 = st.tabs(["Direct Matches", "PLM Not in SAP", "SAP Not in PLM", "70% or more Matches"])
        with tab1:
            st.dataframe(direct_matches)
        with tab2:
            st.dataframe(missing_in_sap)
        with tab3:
            st.dataframe(missing_in_plm)
        with tab4:
            st.dataframe(fuzzy_df)

    except Exception as e:
        st.error(f"❌ Error while processing: {e}")
else:
    st.info("⬆️ Please upload both SAP and PLM Excel files to begin.")




